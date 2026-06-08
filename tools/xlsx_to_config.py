#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_to_config.py —— 把 Excel 点表(.xlsx) 转换成 S7→OPC UA 网关能读的 JSON 配置。

这是一个【独立程序】：只做「读 xlsx → 吐 JSON」，不依赖、也不修改 C 网关。

支持两种点表格式，靠表头自动识别（也可用 --format 强制）：

  ① 中文工业点表(cn)   —— 现场常见的样子，表头是中文：
       点位ID | 点位名称 | 点位地址 | 数据类型 | NodeID | 单位 | 最新数据 | 更新时间 | 设备号
     地址用西门子写法(DB1.DBW2)，类型写 float32 这种，每行带固定 NodeID(ns=2;s=[...])。
     这种表里没有 PLC 的 IP/端口，所以连接信息用命令行参数给(--ip 等)。一张表 = 一台 PLC。

  ② 显式列点表(simple) —— 每个字段一列，直接对应网关配置：
       plc_name ip port rack slot poll_interval_ms tag_name area db start bit type

用法：
  # 中文点表(必须用 --ip 给出 PLC 地址)
  python3 tools/xlsx_to_config.py 点表.xlsx -o out.json --ip 192.168.0.10 --plc-name 炉区PLC
  # 显式列点表
  python3 tools/xlsx_to_config.py 点表.xlsx -o out.json
  # 不给 -o 就打印到屏幕

退出码：
  0 = 成功
  1 = 点表"内容"有错误(逐条列出来，改了重跑即可)
  2 = 根本没法处理(缺列 / 文件打不开 / 没装 openpyxl / sheet 名写错 / 参数非法 / 格式认不出 等)
"""

import argparse
import json
import re
import sys


def die(msg, code=2):
    """打印错误到 stderr 后退出。默认码 2 = 用法/环境/读文件类错误(区别于"点表内容错误"的码 1)。"""
    print(msg, file=sys.stderr)
    sys.exit(code)


# ─────────────────────────────────────────────────────────────────────────────
# 1. 这些"常量"必须和 C 网关里的定义保持一致(改 C 端时记得同步改这里)
# ─────────────────────────────────────────────────────────────────────────────

# C 端 s7_area_from_str 接受的存储区
VALID_AREAS = {"DB", "M", "I", "Q"}

# C 端 s7_type_from_str 接受的 12 种数据类型
VALID_TYPES = {
    "BOOL", "INT", "DINT", "REAL", "BYTE", "SINT",
    "USINT", "WORD", "UINT", "DWORD", "UDINT", "LREAL",
}

# 各类型占几个字节(对齐 C 端 s7_type_size)，用于地址宽度的一致性提示
TYPE_BYTES = {
    "BOOL": 1, "BYTE": 1, "SINT": 1, "USINT": 1,
    "INT": 2, "WORD": 2, "UINT": 2,
    "DINT": 4, "REAL": 4, "DWORD": 4, "UDINT": 4,
    "LREAL": 8,
}

# C 端结构体里的缓冲区长度(config.c 用 strcmp/strlen 校验，按 UTF-8 字节数算)
MAX_PLC_NAME_BYTES = 64    # PlcCfg.name[64]
MAX_IP_BYTES = 16          # PlcCfg.ip[16]
MAX_TAG_NAME_BYTES = 64    # TagCfg.name[64]
MAX_NODE_ID_BYTES = 128    # TagCfg.node_id[128]

# —— 显式列模式(simple)的列名 ——
SIMPLE_REQUIRED = {"plc_name", "ip", "tag_name", "area", "start", "type"}

# —— 中文模式(cn)的列名(都用"小写、去空格后"的形式比对；中文 lower() 不变) ——
CN_ID = "点位id"
CN_NAME = "点位名称"
CN_ADDR = "点位地址"
CN_TYPE = "数据类型"
CN_NODEID = "nodeid"

# 中文/通用 数据类型名 → 网关类型(小写匹配)
CN_TYPE_MAP = {
    "bool": "BOOL", "bit": "BOOL", "boolean": "BOOL",
    "sint": "SINT", "int8": "SINT",
    "byte": "BYTE", "usint": "BYTE", "uint8": "BYTE",
    "int": "INT", "int16": "INT", "short": "INT",
    "word": "WORD", "uint": "WORD", "uint16": "WORD", "ushort": "WORD",
    "dint": "DINT", "int32": "DINT", "long": "DINT",
    "dword": "DWORD", "udint": "DWORD", "uint32": "DWORD", "ulong": "DWORD",
    "real": "REAL", "float": "REAL", "float32": "REAL",
    "lreal": "LREAL", "double": "LREAL", "float64": "LREAL",
}


# ─────────────────────────────────────────────────────────────────────────────
# 2. 小工具：把单元格的值安全地转成"字符串"或"整数"
# ─────────────────────────────────────────────────────────────────────────────

def cell_is_empty(v):
    """空单元格判定：None 或 去掉空格后是空串。"""
    return v is None or (isinstance(v, str) and v.strip() == "")


def to_str(v):
    """把单元格转成去掉首尾空格的字符串；空单元格返回空串 ''。"""
    if cell_is_empty(v):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))      # 1001.0 -> "1001"，别带小数点
    return str(v).strip()


def to_int(v, field, rownum, errors, required, default=None):
    """把单元格转成 int；空/非整数按 required 决定报错还是给默认值。错误收集进 errors。"""
    if cell_is_empty(v):
        if required:
            errors.append(f"第{rownum}行：缺少必填整数列 '{field}'")
            return None
        return default
    try:
        if isinstance(v, bool):
            raise ValueError
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            if not v.is_integer():
                raise ValueError
            return int(v)
        f = float(str(v).strip())
        if not f.is_integer():
            raise ValueError
        return int(f)
    except (ValueError, TypeError):
        errors.append(f"第{rownum}行：列 '{field}' 必须是整数，当前是 {v!r}")
        return None


def check_tag_common(rownum, name, area, typ, db, start, bit, node_id, errors):
    """两种模式共用的 tag 字段校验(对齐 C 端 config.c)。问题收集进 errors。"""
    if area and area not in VALID_AREAS:
        errors.append(f"第{rownum}行：area 非法 '{area}'，只能是 {'/'.join(sorted(VALID_AREAS))}")
    if typ and typ not in VALID_TYPES:
        errors.append(f"第{rownum}行：type 非法 '{typ}'，只能是 {'/'.join(sorted(VALID_TYPES))}")
    if start is not None and start < 0:
        errors.append(f"第{rownum}行：start 不能为负({start})")
    if area == "DB":
        if db is None:
            errors.append(f"第{rownum}行：area=DB 时必须有 db")
        elif db < 1:
            errors.append(f"第{rownum}行：db 必须 >=1(当前 {db})")
    if typ == "BOOL":
        if bit is None:
            errors.append(f"第{rownum}行：type=BOOL 时必须有 bit(0~7)")
        elif bit < 0 or bit > 7:
            errors.append(f"第{rownum}行：bit 必须在 0~7(当前 {bit})")
    if name and len(name.encode("utf-8")) >= MAX_TAG_NAME_BYTES:
        errors.append(f"第{rownum}行：点名太长(C 端上限 {MAX_TAG_NAME_BYTES-1} 字节)")
    if node_id and len(node_id.encode("utf-8")) >= MAX_NODE_ID_BYTES:
        errors.append(f"第{rownum}行：NodeID 太长(C 端上限 {MAX_NODE_ID_BYTES-1} 字节)")


# ─────────────────────────────────────────────────────────────────────────────
# 3. 读取 xlsx：返回 (表头映射, 数据行)。表头映射 = {规范列名(小写去空格) -> 列下标}
#    这里不做"必填列"检查——交给各模式自己查，因为两种格式必填列不同。
# ─────────────────────────────────────────────────────────────────────────────

def load_rows(path, sheet_name):
    try:
        import openpyxl
    except ImportError:
        die("错误：没装 openpyxl。请先运行：python3 -m pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        die(f"错误：找不到文件 {path}")
    except Exception as e:
        die(f"错误：打不开 xlsx 文件 {path}：{e}")

    if sheet_name and sheet_name not in wb.sheetnames:
        die(f"错误：找不到名为 '{sheet_name}' 的工作表。\n"
            f"  这个文件里有这些表：" + "、".join(wb.sheetnames))
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        die("错误：工作簿里没有可用的工作表")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        die("错误：表是空的")

    # 第一行当表头：列名统一"小写、去空格"(中文不受 lower 影响)，建 {列名 -> 下标}
    header_map = {}
    for idx, name in enumerate(all_rows[0]):
        if name is None:
            continue
        key = str(name).strip().lower()
        if key and key not in header_map:
            header_map[key] = idx

    # 其余是数据行；连同"原始行号"返回(行号从 1 数，表头是第 1 行)；跳过整行空白
    data_rows = []
    for i, row in enumerate(all_rows[1:], start=2):
        if all(cell_is_empty(c) for c in row):
            continue
        data_rows.append((i, row))
    return header_map, data_rows


# ─────────────────────────────────────────────────────────────────────────────
# 4. 显式列模式(simple)：每个字段一列，按 plc_name 分组
# ─────────────────────────────────────────────────────────────────────────────

def build_plcs_simple(header_map, data_rows):
    missing = SIMPLE_REQUIRED - set(header_map.keys())
    if missing:
        die("错误：点表缺少必须的列：" + "、".join(sorted(missing)) +
            "\n  显式列模式的必填列有：" + "、".join(sorted(SIMPLE_REQUIRED)))

    errors = []
    plcs = {}
    plc_order = []

    def col(row, name):
        idx = header_map.get(name)
        return row[idx] if (idx is not None and idx < len(row)) else None

    for rownum, row in data_rows:
        plc_name = to_str(col(row, "plc_name"))
        ip = to_str(col(row, "ip"))
        if not plc_name:
            errors.append(f"第{rownum}行：plc_name 不能为空")
        if not ip:
            errors.append(f"第{rownum}行：ip 不能为空")

        port = to_int(col(row, "port"), "port", rownum, errors, required=False, default=102)
        rack = to_int(col(row, "rack"), "rack", rownum, errors, required=False, default=0)
        slot = to_int(col(row, "slot"), "slot", rownum, errors, required=False, default=1)
        poll = to_int(col(row, "poll_interval_ms"), "poll_interval_ms",
                      rownum, errors, required=False, default=None)

        tag_name = to_str(col(row, "tag_name"))
        area = to_str(col(row, "area")).upper()
        typ = to_str(col(row, "type")).upper()
        start = to_int(col(row, "start"), "start", rownum, errors, required=True)
        db = to_int(col(row, "db"), "db", rownum, errors, required=False, default=None)
        bit = to_int(col(row, "bit"), "bit", rownum, errors, required=False, default=None)

        if not tag_name:
            errors.append(f"第{rownum}行：tag_name 不能为空")
        check_tag_common(rownum, tag_name, area, typ, db, start, bit, "", errors)
        if plc_name and len(plc_name.encode("utf-8")) >= MAX_PLC_NAME_BYTES:
            errors.append(f"第{rownum}行：plc_name 太长(C 端上限 {MAX_PLC_NAME_BYTES-1} 字节)")
        if ip and len(ip.encode("utf-8")) >= MAX_IP_BYTES:
            errors.append(f"第{rownum}行：ip 太长(C 端上限 {MAX_IP_BYTES-1} 字节)")
        if port is not None and (port < 1 or port > 65535):
            errors.append(f"第{rownum}行：port 必须在 1~65535(当前 {port})")
        if poll is not None and poll <= 0:
            errors.append(f"第{rownum}行：poll_interval_ms 必须 >0(当前 {poll})")

        if not plc_name or not ip or not tag_name:
            continue

        conn = {"ip": ip, "port": port, "rack": rack, "slot": slot}
        if plc_name not in plcs:
            plc = {"name": plc_name, **conn}
            if poll is not None:
                plc["poll_interval_ms"] = poll
            plc["_poll_raw"] = poll
            plc["tags"] = []
            plc["_tagnames"] = set()
            plcs[plc_name] = plc
            plc_order.append(plc_name)
        else:
            plc = plcs[plc_name]
            prev_conn = {k: plc[k] for k in ("ip", "port", "rack", "slot")}
            if prev_conn != conn:
                errors.append(
                    f"第{rownum}行：PLC '{plc_name}' 的连接参数和前面不一致"
                    f"(之前 {prev_conn}，这里 {conn})——同一个 PLC 名必须用同样的 ip/port/rack/slot")
            if poll != plc["_poll_raw"]:
                errors.append(
                    f"第{rownum}行：PLC '{plc_name}' 的 poll_interval_ms 和前面不一致"
                    f"(之前 {plc['_poll_raw']}，这里 {poll})——同一个 PLC 要么都不填、要么填同一个值")

        if tag_name in plc["_tagnames"]:
            errors.append(f"第{rownum}行：PLC '{plc_name}' 里 tag 名 '{tag_name}' 重复了")
            continue
        plc["_tagnames"].add(tag_name)

        tag = {"name": tag_name, "area": area}
        if area == "DB":
            tag["db"] = db
        tag["start"] = start
        if typ == "BOOL":
            tag["bit"] = bit
        tag["type"] = typ
        plc["tags"].append(tag)

    plcs_list = []
    for name in plc_order:
        plc = plcs[name]
        plc.pop("_tagnames", None)
        plc.pop("_poll_raw", None)
        plcs_list.append(plc)

    return plcs_list, errors, []


# ─────────────────────────────────────────────────────────────────────────────
# 5. 中文工业点表模式(cn)：解析西门子地址 + 类型映射 + NodeID 透传
# ─────────────────────────────────────────────────────────────────────────────

# 西门子绝对地址写法。X=位, B=字节, W=字(2字节), D=双字(4字节)
_RE_DB_BIT = re.compile(r"^DB(\d+)\.DBX(\d+)\.([0-7])$")   # DB1.DBX8.0
_RE_DB_NUM = re.compile(r"^DB(\d+)\.DB([BWD])(\d+)$")        # DB1.DBW2
_RE_MIQ_BIT = re.compile(r"^([MIQ])(\d+)\.([0-7])$")         # M10.0
_RE_MIQ_NUM = re.compile(r"^([MIQ])([BWD])(\d+)$")           # MW10


def parse_s7_address(addr):
    """
    解析西门子地址，返回 (area, db, start, bit, width_letter)。
      - db   非 DB 区为 None
      - bit  非位地址为 None
      - width_letter ∈ {X,B,W,D}，用于和数据类型做宽度一致性提示
    解析不了抛 ValueError。
    """
    s = addr.upper().replace(" ", "")
    m = _RE_DB_BIT.match(s)
    if m:
        return ("DB", int(m.group(1)), int(m.group(2)), int(m.group(3)), "X")
    m = _RE_DB_NUM.match(s)
    if m:
        return ("DB", int(m.group(1)), int(m.group(3)), None, m.group(2))
    m = _RE_MIQ_BIT.match(s)
    if m:
        return (m.group(1), None, int(m.group(2)), int(m.group(3)), "X")
    m = _RE_MIQ_NUM.match(s)
    if m:
        return (m.group(1), None, int(m.group(3)), None, m.group(2))
    raise ValueError(f"无法解析点位地址 '{addr}'(支持 DB1.DBW2 / DB1.DBX0.0 / MW10 / M10.0 这类写法)")


def build_plcs_cn(header_map, data_rows, conn):
    """conn = {plc_name, ip, port, rack, slot}（来自命令行）。返回 (plcs, errors, warnings)。"""
    # —— 结构性检查：必须有的列(缺了直接退出码 2) ——
    has_name = CN_NAME in header_map
    has_id = CN_ID in header_map
    if CN_ADDR not in header_map:
        die("错误：中文点表缺少『点位地址』列")
    if CN_TYPE not in header_map:
        die("错误：中文点表缺少『数据类型』列")
    if not has_name and not has_id:
        die("错误：中文点表至少要有『点位名称』或『点位ID』列之一(用来给点命名)")
    has_nodeid = CN_NODEID in header_map

    errors = []
    warnings = []
    if not has_nodeid:
        warnings.append("点表里没有『NodeID』列——所有点都会由网关自动生成 NodeID(ns=1;s=PLC.点名)")

    tags = []
    seen = set()

    def col(row, name):
        idx = header_map.get(name)
        return row[idx] if (idx is not None and idx < len(row)) else None

    for rownum, row in data_rows:
        name = to_str(col(row, CN_NAME)) if has_name else ""
        if not name and has_id:
            name = to_str(col(row, CN_ID))          # 没名字就拿点位ID当名字
        addr = to_str(col(row, CN_ADDR))
        raw_type = to_str(col(row, CN_TYPE))
        node_id = to_str(col(row, CN_NODEID)) if has_nodeid else ""

        if not name:
            errors.append(f"第{rownum}行：点位名称和点位ID都空，没法命名")
            continue
        if not addr:
            errors.append(f"第{rownum}行：点位地址为空")
            continue

        # 解析地址
        try:
            area, db, start, bit, width = parse_s7_address(addr)
        except ValueError as e:
            errors.append(f"第{rownum}行：{e}")
            continue

        # 映射数据类型
        typ = CN_TYPE_MAP.get(raw_type.lower())
        if typ is None:
            errors.append(f"第{rownum}行：数据类型 '{raw_type}' 不认识"
                          f"(支持 bool/int16/int32/word/dword/float32/float64 等)")
            continue

        # 位地址 ↔ BOOL 必须对应
        if typ == "BOOL" and bit is None:
            errors.append(f"第{rownum}行：类型是 BOOL，但地址 '{addr}' 不是位地址"
                          f"(应形如 DB1.DBX0.0 或 M0.0)")
            continue
        if typ != "BOOL" and bit is not None:
            errors.append(f"第{rownum}行：地址 '{addr}' 是位地址，但类型 '{typ}' 不是 BOOL")
            continue

        # 地址宽度字母和类型字节数不符：只提示不拦(以类型为准)
        if width in ("B", "W", "D"):
            want = {"B": 1, "W": 2, "D": 4}[width]
            got = TYPE_BYTES.get(typ, want)
            if want != got:
                hint = {1: "DBB/MB", 2: "DBW/MW", 4: "DBD/MD"}.get(got, "")
                msg = (f"第{rownum}行：地址 '{addr}' 宽度 {width}={want}字节，"
                       f"但类型 {typ} 占 {got}字节——按类型取 {got} 字节")
                if hint:
                    msg += f"(地道写法宽度字母该用 {hint} 那种)"
                warnings.append(msg)

        # 共用校验(长度/范围/DB需db/BOOL需bit)
        check_tag_common(rownum, name, area, typ, db, start, bit, node_id, errors)

        if has_nodeid and not node_id:
            warnings.append(f"第{rownum}行：NodeID 为空，这个点会由网关自动生成 NodeID")

        if name in seen:
            errors.append(f"第{rownum}行：点位名 '{name}' 重复了(同一台 PLC 里点名不能重)")
            continue
        seen.add(name)

        tag = {"name": name, "area": area}
        if area == "DB":
            tag["db"] = db
        tag["start"] = start
        if typ == "BOOL":
            tag["bit"] = bit
        tag["type"] = typ
        if node_id:
            tag["node_id"] = node_id        # 透传点表里的 NodeID(网关会照它暴露)
        tags.append(tag)

    plc = {
        "name": conn["plc_name"], "ip": conn["ip"], "port": conn["port"],
        "rack": conn["rack"], "slot": conn["slot"], "tags": tags,
    }
    return [plc], errors, warnings


# ─────────────────────────────────────────────────────────────────────────────
# 6. 格式识别 + 主流程
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(header_map, override):
    if override and override != "auto":
        return override
    if CN_ADDR in header_map:           # 有"点位地址"列 -> 中文工业点表
        return "cn"
    if "area" in header_map and "start" in header_map:
        return "simple"
    return None


def main():
    ap = argparse.ArgumentParser(
        description="把 Excel 点表(.xlsx) 转成 S7→OPC UA 网关用的 JSON 配置。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("input", help="输入的 xlsx 点表路径")
    ap.add_argument("-o", "--output", help="输出 JSON 路径(不给就打印到屏幕)")
    ap.add_argument("--format", choices=["auto", "cn", "simple"], default="auto",
                    help="点表格式(默认 auto 自动识别)")
    # 全局
    ap.add_argument("--opcua-port", type=int, default=4840, help="OPC UA 服务端口(默认 4840)")
    ap.add_argument("--cache-ttl-ms", type=int, default=1000, help="缓存有效期毫秒(默认 1000)")
    ap.add_argument("--sheet", default=None, help="指定工作表名(默认第一张)")
    # 中文模式专用：连接信息(点表里没有)
    ap.add_argument("--ip", default=None, help="[中文模式必填] PLC 的 IP")
    ap.add_argument("--port", type=int, default=102, help="[中文模式] PLC 端口(默认 102)")
    ap.add_argument("--rack", type=int, default=0, help="[中文模式] 机架号(默认 0)")
    ap.add_argument("--slot", type=int, default=1, help="[中文模式] 槽位号(默认 1)")
    ap.add_argument("--plc-name", default="PLC1", help="[中文模式] PLC 名(默认 PLC1)")
    args = ap.parse_args()

    if args.opcua_port < 1 or args.opcua_port > 65535:
        die(f"错误：--opcua-port 必须在 1~65535(当前 {args.opcua_port})")
    if args.cache_ttl_ms <= 0:
        die(f"错误：--cache-ttl-ms 必须 >0(当前 {args.cache_ttl_ms})")

    header_map, data_rows = load_rows(args.input, args.sheet)
    if not data_rows:
        die("错误：点表里没有数据行(除了表头)")

    fmt = detect_format(header_map, args.format)
    if fmt is None:
        die("错误：认不出点表格式。\n"
            "  中文模式需要『点位地址』列；显式列模式需要 area + start 列。\n"
            "  也可以用 --format cn / --format simple 强制指定。")

    if fmt == "cn":
        if not args.ip:
            die("错误：中文点表模式需要用 --ip 指定 PLC 的 IP(点表里没有这个信息)。\n"
                "  例：--ip 192.168.0.10 --port 102 --rack 0 --slot 1 --plc-name 炉区PLC")
        if args.port < 1 or args.port > 65535:
            die(f"错误：--port 必须在 1~65535(当前 {args.port})")
        if args.rack < 0 or args.slot < 0:
            die("错误：--rack/--slot 不能为负")
        if len(args.plc_name.encode("utf-8")) >= MAX_PLC_NAME_BYTES:
            die(f"错误：--plc-name 太长(上限 {MAX_PLC_NAME_BYTES-1} 字节)")
        if len(args.ip.encode("utf-8")) >= MAX_IP_BYTES:
            die(f"错误：--ip 太长(上限 {MAX_IP_BYTES-1} 字节)")
        conn = {"plc_name": args.plc_name, "ip": args.ip, "port": args.port,
                "rack": args.rack, "slot": args.slot}
        plcs, errors, warnings = build_plcs_cn(header_map, data_rows, conn)
    else:
        plcs, errors, warnings = build_plcs_simple(header_map, data_rows)

    if errors:
        print(f"❌ 点表有 {len(errors)} 个问题，没有生成配置：", file=sys.stderr)
        for e in errors:
            print("   - " + e, file=sys.stderr)
        sys.exit(1)

    if not plcs or all(len(p["tags"]) == 0 for p in plcs):
        die("错误：没有解析出任何点位")

    config = {
        "opcua": {"port": args.opcua_port},
        "collection": {"cache_ttl_ms": args.cache_ttl_ms},
        "plcs": plcs,
    }
    text = json.dumps(config, indent=2, ensure_ascii=False)

    if args.output:
        try:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(text + "\n")
        except OSError as e:
            die(f"错误：写不了输出文件 {args.output}：{e}")
        total_tags = sum(len(p["tags"]) for p in plcs)
        print(f"✅ 已生成 {args.output}（格式：{fmt}）")
        print(f"   共 {len(plcs)} 个 PLC、{total_tags} 个点位：")
        for p in plcs:
            n_nid = sum(1 for t in p["tags"] if "node_id" in t)
            extra = f"，{n_nid} 个用点表 NodeID" if n_nid else ""
            print(f"     - {p['name']}  {p['ip']}:{p['port']}  ({len(p['tags'])} 点{extra})")
    else:
        print(text)

    # 提示(非致命)放最后，免得被成功信息淹没
    if warnings:
        print(f"\n⚠️  {len(warnings)} 条提示：", file=sys.stderr)
        for w in warnings:
            print("   - " + w, file=sys.stderr)


if __name__ == "__main__":
    main()
